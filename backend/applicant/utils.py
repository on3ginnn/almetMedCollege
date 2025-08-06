from docxtpl import DocxTemplate
from django.http import HttpResponse
from django.template.defaultfilters import date as _date
import io
from datetime import datetime
from .models import Applicant
from io import BytesIO
from collections import defaultdict
import openpyxl
from django.db import models
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.utils.timezone import localtime

MIN_COLUMNS_WIDTH_EXCEL_SHEET = 20


def generate_applicants_rating(applicants):
    try:
        # Фильтруем только тех, кто забрал документы (canceled=True)

        # Создаём книгу
        wb = openpyxl.Workbook()
        ws = wb.active

        # Список всех полей модели (кроме служебных и ненужных)
        exclude_fields = ["id", 'enrolled', 'documents_delivered', 'documents_canceled']
        fields = [f.name for f in Applicant._meta.fields if f.name not in exclude_fields]

        # Заголовки для Excel (читаемые названия)
        headers = [Applicant._meta.get_field(f).verbose_name for f in fields]
        ws.append(headers)

        # Добавляем данные
        for applicant in applicants:
            row_data = []
            for field in fields:
                # Если у поля есть display-метод (для choices), используем его
                display_method = f"get_{field}_display"
                if hasattr(applicant, display_method):
                    value = getattr(applicant, display_method)()
                else:
                    value = getattr(applicant, field)

                # Обрабатываем типы данных
                if value is None:
                    value = ""
                elif isinstance(value, bool):
                    value = "Да" if value else "Нет"
                elif hasattr(value, 'strftime'):  # Date, DateTime
                    value = value.strftime('%d.%m.%Y')
                row_data.append(value)
            ws.append(row_data)

        # Выравниваем текст по верхнему краю
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Устанавливаем ширину колонок
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in list(col_cells)[1:])
            adjusted_width = min(max(max_length + 2, MIN_COLUMNS_WIDTH_EXCEL_SHEET), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Сохраняем книгу в поток
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
    except Exception as e:
        print(e)


def generate_applicants_canceled_excel(applicants):
    try:
        # Фильтруем только тех, кто забрал документы (canceled=True)

        # Создаём книгу
        wb = openpyxl.Workbook()
        ws = wb.active

        # Список всех полей модели (кроме служебных и ненужных)
        exclude_fields = ["id", 'enrolled', 'documents_delivered', 'documents_canceled']
        fields = [f.name for f in Applicant._meta.fields if f.name not in exclude_fields]

        # Заголовки для Excel (читаемые названия)
        headers = [Applicant._meta.get_field(f).verbose_name for f in fields]
        ws.append(headers)

        # Добавляем данные
        for applicant in applicants:
            row_data = []
            for field in fields:
                # Если у поля есть display-метод (для choices), используем его
                display_method = f"get_{field}_display"
                if hasattr(applicant, display_method):
                    value = getattr(applicant, display_method)()
                else:
                    value = getattr(applicant, field)

                # Обрабатываем типы данных
                if value is None:
                    value = ""
                elif isinstance(value, bool):
                    value = "Да" if value else "Нет"
                elif hasattr(value, 'strftime'):  # Date, DateTime
                    value = value.strftime('%d.%m.%Y')
                row_data.append(value)
            ws.append(row_data)

        # Выравниваем текст по верхнему краю
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Устанавливаем ширину колонок
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in list(col_cells)[1:])
            adjusted_width = min(max(max_length + 2, MIN_COLUMNS_WIDTH_EXCEL_SHEET), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Сохраняем книгу в поток
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
    except Exception as e:
        print(e)

def generate_applicants_excel(applicants):
    try:
        # Создаём книгу
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Удаляем пустой лист

        # Список всех полей модели (кроме служебных и ненужных)
        exclude_fields = ["id", 'specialty', 'enrolled']
        fields = [f.name for f in Applicant._meta.fields if f.name not in exclude_fields]

        # Заголовки для Excel (читаемые названия)
        headers = [Applicant._meta.get_field(f).verbose_name for f in fields]

        # Проходим по специальностям и создаём отдельные листы
        for spec_key, spec_name in Applicant.SPECIALTY_CHOICES:
            # Фильтруем абитуриентов по специальности
            spec_applicants = applicants.filter(specialty=spec_key)

            # Создаём лист с названием специальности
            ws = wb.create_sheet(title=spec_name[:30])  # Excel ограничивает имя до 31 символа

            # Добавляем заголовки
            ws.append(headers)

            # Добавляем данные
            for applicant in spec_applicants:
                row_data = []
                for field in fields:
                    # Если у поля есть display-метод (для choices), используем его
                    display_method = f"get_{field}_display"
                    if hasattr(applicant, display_method):
                        value = getattr(applicant, display_method)()
                    else:
                        value = getattr(applicant, field)

                    # Обрабатываем типы данных
                    if value is None:
                        value = ""
                    elif isinstance(value, bool):
                        value = "Да" if value else "Нет"
                    elif hasattr(value, 'strftime'):  # Date, DateTime
                        value = value.strftime('%d.%m.%Y')
                    row_data.append(value)
                ws.append(row_data)

            # Выравниваем текст по верхнему краю
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Устанавливаем ширину колонок
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in list(col_cells)[1:])
                adjusted_width = min(max(max_length + 2, MIN_COLUMNS_WIDTH_EXCEL_SHEET), 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Сохраняем книгу в поток
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
    except Exception as e:
        print(e)


def generate_application_docx(applicant: Applicant):
    template_path = 'static/docx/applicant.docx'
    doc = DocxTemplate(template_path)

    # Split full_name into parts
    name_parts = applicant.full_name.split() if applicant.full_name else ['', '', '']
    last_name = name_parts[0] if len(name_parts) > 0 else ''
    first_name = name_parts[1] if len(name_parts) > 1 else ''
    father_name = name_parts[2] if len(name_parts) > 2 else ''

    # Map specialty to human-readable form
    # specialty_map = {
    #     'pharmacy': 'Фармация',
    #     'nursing': f'Сестринское дело ({dict(Applicant.EDUCATION_BASE_CHOICES).get(applicant.education_base, "")})',
    #     'midwifery': 'Акушерское дело',
    #     'lab_diagnostics': 'Лабораторная диагностика',
    #     'medical_treatment': 'Лечебное дело',
    # }
    # specialty = specialty_map.get(applicant.specialty, '')
    specialty = dict(Applicant.SPECIALTY_CHOICES).get(applicant.specialty, '')

    # Clean SNILS
    clean_snils = applicant.snils.replace('-', '').replace(' ', '') if applicant.snils else ''

    # Format representative passport details
    representative1_passport = []
    if applicant.representative1_passport_series:
        representative1_passport.append(f"Серия: {applicant.representative1_passport_series}")
    if applicant.representative1_passport_number:
        representative1_passport.append(f"№ {applicant.representative1_passport_number}")
    if applicant.representative1_passport_issued_by:
        representative1_passport.append(f"выдан: {applicant.representative1_passport_issued_by}")
    if applicant.representative1_passport_issued_date:
        representative1_passport.append(_date(applicant.representative1_passport_issued_date, "d.m.Y"))

    representative1_passport = ', '.join(representative1_passport) if representative1_passport else '''______________________________________________
    _________________________________________________________________________________________
    '''

    representative2_passport = []
    if applicant.representative2_passport_series:
        representative2_passport.append(f"Серия: {applicant.representative2_passport_series}")
    if applicant.representative2_passport_number:
        representative2_passport.append(f"№ {applicant.representative2_passport_number}")
    if applicant.representative2_passport_issued_by:
        representative2_passport.append(f"выдан: {applicant.representative2_passport_issued_by}")
    if applicant.representative2_passport_issued_date:
        representative2_passport.append(_date(applicant.representative2_passport_issued_date, "d.m.Y"))

    representative2_passport = ', '.join(representative2_passport) if representative2_passport else '''______________________________________________
    _________________________________________________________________________________________
    '''

    print(representative2_passport)

    # Determine social benefits
    benefits_enrollment = 'Нет' if applicant.priority_enrollment == 'none' and applicant.preferential_enrollment == 'none' else 'Да'

    # Calculate average grade dynamically
    grades = [grade for grade in [applicant.grade_russian, applicant.grade_biology, applicant.grade_chemistry] if grade is not None]
    average_grade = sum(grades) / len(grades) if grades else ''

    # Context for template rendering
    context = {
        'registration_number': applicant.registration_number or '_____________________',
        'last_name': last_name,
        'first_name': first_name,
        'father_name': father_name or '__________________',
        'birth_date': _date(applicant.birth_date, 'd.m.Y') if applicant.birth_date else '',
        'birth_place': applicant.birth_place or '''_______________________
______________________________________
''',
        'citizenship': applicant.citizenship or '___________________________',
        'passport_series': applicant.passport_series or '______',
        'passport_number': applicant.passport_number or '_____________',
        'passport_issued_date': _date(applicant.passport_issued_date, 'd.m.Y') if applicant.passport_issued_date else '___________________________',
        'passport_issued_by': applicant.passport_issued_by or '''_____________________________
_______________________________________

''',
        'passport_division_code': applicant.passport_division_code or '__________',
        'address': applicant.address or '_____________________________________________________________________________________',
        'address_actual': applicant.address_actual or applicant.address or 'тот же',
        'passport_registration_date': _date(applicant.passport_registration_date, 'd.m.Y') if applicant.passport_registration_date else '',
        'snils': clean_snils,
        'inn': applicant.inn or '___________________',
        'medical_policy': applicant.medical_policy or '',
        'military_id': 'Да' if applicant.military_id else 'Нет',
        'student_phone': applicant.student_phone or '____________________',
        'email': applicant.student_email or '______________________________________',
        'specialty': specialty,
        'study_form_verbose': dict(Applicant.STUDY_FORM_CHOICES).get(applicant.study_form, ''),
        'documents_submitted': applicant.documents_submitted or '',
        'admission_type': dict(Applicant.ADMISSION_TYPE_CHOICES).get(applicant.admission_type, ''),
        'graduation_year': str(applicant.graduation_year) if applicant.graduation_year else '__________',
        'graduation_institution': applicant.graduation_institution or '''__________________________________________________________________________________________________________________________________________________________________________''',
        'certificate_series': applicant.certificate_series or '___________________',
        'certificate_issued_date': _date(applicant.certificate_issued_date, 'd.m.Y') if applicant.certificate_issued_date else '_________________________',
        # 'average_grade': str(average_grade) if average_grade else '_________________________',
        'average_grade': str(applicant.average_grade),
        'grade_russian': str(applicant.grade_russian) if applicant.grade_russian else '___________________',
        'grade_biology': str(applicant.grade_biology) if applicant.grade_biology else '_________________________',
        'grade_chemistry': str(applicant.grade_chemistry) if applicant.grade_chemistry else '____________________________',
        'benefits_enrollment': benefits_enrollment,
        'needs_dormitory': 'нуждаюсь' if applicant.needs_dormitory else 'не нуждаюсь',
        'representative1_fio': applicant.representative1_name or '''________________________________________________________
_________________________________________________________________________________________
''',
        'representative1_job': applicant.representative1_job or '___________________________________________________________________',
        'representative1_phone': applicant.representative1_phone or '_______________________________________________________________________',
        'representative1_passport': representative1_passport,
        'representative2_fio': applicant.representative2_name or '___________________________________________________________',
        'representative2_job': applicant.representative2_job or '___________________________________________________________________',
        'representative2_phone': applicant.representative2_phone or '________________________________________________________________________',
        'representative2_passport': representative2_passport,
        'current_date': _date(datetime.now(), 'd F Y'),
        'applicant_signature': '',
        'commission_signature': '',
    }

    # Render the template with the context
    doc.render(context)

    buffer = io.BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    return buffer


def generate_application_titul(applicant: Applicant):
    template_path = 'static/docx/applicant_titul.docx'
    doc = DocxTemplate(template_path)

    # Карта специальностей
    # specialty_map = {
    #     'pharmacy': 'Фармация',
    #     'nursing': f'Сестринское дело ({dict(Applicant.EDUCATION_BASE_CHOICES).get(applicant.education_base, "")})',
    #     'midwifery': 'Акушерское дело',
    #     'lab_diagnostics': 'Лабораторная диагностика',
    #     'medical_treatment': 'Лечебное дело',
    # }

    # specialty = specialty_map.get(applicant.specialty, '')
    specialty = dict(Applicant.SPECIALTY_CHOICES).get(applicant.specialty, '')
    clean_snils = applicant.snils.replace('-', '').replace(' ', '') if applicant.snils else ''

    # Представители
    mother_info = f"{applicant.representative1_name or ''}, {applicant.representative1_phone or ''}".strip(', ')
    father_info = f"{applicant.representative2_name or ''}, {applicant.representative2_phone or ''}".strip(', ')

    # Паспорт
    passport_info = ''
    if applicant.passport_series and applicant.passport_number:
        passport_info += f"Серия: {applicant.passport_series} № {applicant.passport_number}, "
    if applicant.passport_issued_by:
        passport_info += f"{applicant.passport_issued_by}, "
    if applicant.passport_issued_date:
        passport_info += f"{_date(applicant.passport_issued_date, 'd.m.Y')}"

    # Дата окончания + учреждение
    graduation_info = ''
    if applicant.graduation_year or applicant.graduation_institution:
        graduation_info = f"{applicant.graduation_year or ''}, {applicant.graduation_institution or ''}".strip(', ')

    # Средний балл
    grades = [g for g in [applicant.grade_russian, applicant.grade_biology, applicant.grade_chemistry] if g is not None]
    average_grade = f"{(sum(grades) / len(grades)):.2f}" if grades else ''

    # Контекст
    context = {
        'registration_number': applicant.registration_number or '',
        'specialty': specialty,
        'full_name': applicant.full_name,
        'citizenship': applicant.citizenship or '',
        'nationality': applicant.nationality or '',
        'birth_date': _date(applicant.birth_date, 'd.m.Y'),
        'birth_place': applicant.birth_place or '',
        'address_actual': applicant.address_actual,

        'certificate_series': applicant.certificate_series or '',
        'graduation_info': graduation_info,

        'passport': passport_info or '',
        'inn': applicant.inn or '',
        'snils': clean_snils,
        'medical_policy': applicant.medical_policy or '',
        'military_id': 'Да' if applicant.military_id else 'Нет',
        'student_phone': applicant.student_phone or '',

        'mother': mother_info,
        'mother_job': applicant.representative1_job or '',
        'father': father_info,
        'father_job': applicant.representative2_job or '',

        'medical_agreement': 'Да' if applicant.medical_contract else 'Нет',  # оставить пустым
        'grade_russian': str(applicant.grade_russian) if applicant.grade_russian else '',
        'grade_biology': str(applicant.grade_biology) if applicant.grade_biology else '',
        'grade_chemistry': str(applicant.grade_chemistry) if applicant.grade_chemistry else '',
        'grade_math': str(applicant.grade_math) if applicant.grade_math else '',
        'grade_foreign': str(applicant.grade_language) if applicant.grade_language else '',
        'grade_physics': str(applicant.grade_physics) if applicant.grade_physics else '',
        # 'average_grade': average_grade,
        'average_grade': str(applicant.average_grade),
    }

    # Рендер
    doc.render(context)
    buffer = io.BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    return buffer
